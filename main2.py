import csv
import math
import multiprocessing

from openpyxl import load_workbook


def calculate_distance(lat1, lon1, lat2, lon2):
    """
    Calculate the distance between two coordinates using the Haversine formula.
    """
    R = 6371  # Radius of the Earth in kilometers
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat / 2) * math.sin(dlat / 2) + math.cos(
        math.radians(lat1)
    ) * math.cos(math.radians(lat2)) * math.sin(dlon / 2) * math.sin(dlon / 2)
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    distance = R * c
    return distance


def calculate_distances(coord):
    """
    Calculate the distances from the source coordinates to each target coordinate.
    """
    utranCell, azimuth, hbw, lat, lon, sheet = coord
    distances = []
    lat_idx, lon_idx = (
        4,
        5,
    )  # Assuming Latitude and Longitude are in the 5th and 6th columns

    for row in sheet.iter_rows(min_row=2, values_only=True):
        target_azimuth = row[1]  # Assuming the azimuth is in the second column
        if target_azimuth == azimuth:
            target_lat = row[lat_idx - 1]
            target_lon = row[lon_idx - 1]
            distance = calculate_distance(lat, lon, target_lat, target_lon)
            distances.append(
                (utranCell, azimuth, hbw, lat, lon, distance, row[0])
            )  # Include utranCell in the result

    return distances


def find_nearest_distances(source_file):
    """
    Find the 30 nearest distances based on the specified azimuth criteria.
    """
    wb = load_workbook(source_file)
    sheet = wb.active

    # Get the column indices for the required headers
    header_row = sheet[1]
    headers = [cell.value for cell in header_row]

    utranCell_idx = headers.index("utranCell") + 1
    HBW_idx = headers.index("HBW") + 1
    lat_idx = headers.index("Latitude") + 1
    lon_idx = headers.index("Longitude") + 1
    azimuth_idx = headers.index("azimuth") + 1

    # Create a list of source coordinates
    source_coords = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        utranCell = row[utranCell_idx - 1]
        azimuth = row[azimuth_idx - 1]
        hbw = row[HBW_idx - 1]
        lat = row[lat_idx - 1]
        lon = row[lon_idx - 1]
        source_coords.append((utranCell, azimuth, hbw, lat, lon, sheet))

    # Use multiprocessing to speed up the calculations
    pool = multiprocessing.Pool()
    results = pool.map(calculate_distances, source_coords)
    pool.close()
    pool.join()

    # Combine the results from all processes
    distances = [item for sublist in results for item in sublist]

    # Filter out duplicate utranCell values
    unique_distances = {}
    for distance in distances:
        utranCell = distance[0]
        if (
            utranCell not in unique_distances
            or distance[5] < unique_distances[utranCell][5]
        ):
            unique_distances[utranCell] = distance

    # Sort the distances and get the 30 nearest distances
    sorted_distances = sorted(unique_distances.values(), key=lambda x: x[5])[:30]

    # Write the results to a CSV file
    with open("output.csv", "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "utranCell",
                "azimuth",
                "HBW",
                "Latitude",
                "Longitude",
                "Neighbor Distance",
                "Beamwidth",
                "Nearest utranCell",
            ]
        )
        writer.writerows(sorted_distances)

    print("Output saved to output.csv")


if __name__ == "__main__":
    source_file = "source_file.xlsx"
    find_nearest_distances(source_file)
