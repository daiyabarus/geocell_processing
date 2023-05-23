import sys
from functools import partial
from math import atan2, cos, radians, sin, sqrt
from multiprocessing import Pool, cpu_count

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from enumlist.enumlist import Enum, SiteColumnHeader, UMTSColumnHeader


def process_excel_file(source_file):
    wb = openpyxl.load_workbook(source_file, read_only=True)
    umts_cells = []
    sites = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        if sheet_name == "umtsCells":
            process_umts_cells(sheet, umts_cells)
        elif sheet_name == "sites":
            process_sites(sheet, sites)

    return umts_cells, sites


def process_umts_cells(sheet, umts_cells):
    umts_headers = [column.value for column in UMTSColumnHeader]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        umts_row = dict(zip(umts_headers, row))
        umts_cells.append(umts_row)


def process_sites(sheet, sites):
    site_headers = [column.value for column in SiteColumnHeader]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        site_row = dict(zip(site_headers, row))
        sites.append(site_row)


def calculate_distance(lat1, lon1, lat2, lon2):
    R = 6371  # Radius of the Earth in kilometers

    # Convert latitude and longitude to radians
    lat1_rad = radians(lat1)
    lon1_rad = radians(lon1)
    lat2_rad = radians(lat2)
    lon2_rad = radians(lon2)

    # Haversine formula
    dlon = lon2_rad - lon1_rad
    dlat = lat2_rad - lat1_rad
    a = sin(dlat / 2) ** 2 + cos(lat1_rad) * cos(lat2_rad) * sin(dlon / 2) ** 2
    c = 2 * atan2(sqrt(a), sqrt(1 - a))
    distance = R * c

    return distance


def calculate_nearest_distance(umts_cells, sites):
    nearest_distance_data = []

    for umts_row in umts_cells:
        umts_loc_code = umts_row[UMTSColumnHeader.LOC_CODE.value]

        nearest_distance = float("inf")
        nearest_site = None

        for site_row in sites:
            site_loc_code = site_row[SiteColumnHeader.LOC_CODE.value]

            if umts_loc_code == site_loc_code:
                umts_lat = umts_row[UMTSColumnHeader.LATITUDE.value]
                umts_lon = umts_row[UMTSColumnHeader.LONGITUDE.value]
                site_lat = site_row[SiteColumnHeader.LATITUDE.value]
                site_lon = site_row[SiteColumnHeader.LONGITUDE.value]

                distance = calculate_distance(umts_lat, umts_lon, site_lat, site_lon)
                if distance < nearest_distance:
                    nearest_distance = distance
                    nearest_site = site_row

        beam_size = 15 if umts_row[UMTSColumnHeader.FREQUENCY.value] == 850 else 30
        ant_size = nearest_distance / 4

        nearest_distance_data.append(
            {
                **umts_row,
                **nearest_site,
                "Nearest_Distance": nearest_distance,
                "Beam_Size": beam_size,
                "Ant_Size": ant_size,
            }
        )

    return nearest_distance_data
