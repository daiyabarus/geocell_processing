import csv
import sys
from datetime import datetime
from math import atan2, cos, radians, sin, sqrt
from multiprocessing import Pool

import openpyxl


def calculate_nearest_distance(row_umtscells, sheet_sites):
    locCode = row_umtscells[2]

    # Mencari data di sheet sites berdasarkan locCode
    site_data = None
    for row_sites in sheet_sites.iter_rows(min_row=2, values_only=True):
        if row_sites[0] == locCode:
            site_data = row_sites
            break

    # Jika data ditemukan
    if site_data is not None:
        # Mendapatkan data dari umtsCells
        utranCell = row_umtscells[0]
        nodeB = row_umtscells[1]
        cellId = row_umtscells[3]
        sc = row_umtscells[4]
        lac = row_umtscells[5]
        rac = row_umtscells[6]
        uarfcnUl = row_umtscells[7]
        uarfcnDl = row_umtscells[8]
        azimuth = row_umtscells[9]
        HBW = row_umtscells[10]
        erpSectorStatusDesc = row_umtscells[11]
        frequency = row_umtscells[12]

        # Mendapatkan data dari sites
        latitude = site_data[2]
        longitude = site_data[3]
        siteName = site_data[4]
        csd = site_data[5]
        cma = site_data[6]
        er = site_data[7]
        prv = site_data[8]
        towerType = site_data[9]
        rnc = site_data[10]

        # Konversi latitude dan longitude menjadi float
        latitude = float(latitude)
        longitude = float(longitude)

        # Menghitung nearest distance berdasarkan azimuth
        nearest_distance = None
        nearest_lc = None

        for row_sites in sheet_sites.iter_rows(min_row=2, values_only=True):
            lat_ref = float(row_sites[2])
            long_ref = float(row_sites[3])

            # Menghitung jarak menggunakan Haversine formula
            R = 6371  # Radius bumi dalam kilometer
            dlat = radians(lat_ref - latitude)
            dlong = radians(long_ref - longitude)
            a = sin(dlat / 2) * sin(dlat / 2) + cos(radians(latitude)) * cos(
                radians(lat_ref)
            ) * sin(dlong / 2) * sin(dlong / 2)
            c = 2 * atan2(sqrt(a), sqrt(1 - a))
            distance = R * c

            if nearest_distance is None or distance < nearest_distance:
                nearest_distance = distance
                nearest_lc = row_sites[0]

        return [
            utranCell,
            nodeB,
            locCode,
            cellId,
            sc,
            lac,
            rac,
            uarfcnUl,
            uarfcnDl,
            azimuth,
            HBW,
            erpSectorStatusDesc,
            frequency,
            latitude,
            longitude,
            nearest_distance,
            nearest_lc,
            siteName,
            csd,
            cma,
            er,
            prv,
            towerType,
            rnc,
        ]


def main(file_name):
    # Membaca file newReference.xlsx
    wb = openpyxl.load_workbook(file_name)
    sheet_umtscells = wb["umtsCells"]
    sheet_sites = wb["sites"]

    # Membuat header untuk file CSV
    header = [
        "utranCell",
        "nodeB",
        "locCode",
        "cellId",
        "sc",
        "lac",
        "rac",
        "uarfcnUl",
        "uarfcnDl",
        "azimuth",
        "HBW",
        "erpSectorStatusDesc",
        "frequency",
        "latitude",
        "longitude",
        "nearest_distance",
        "nearest_lc",
        "siteName",
        "csd",
        "cma",
        "er",
        "prv",
        "towerType",
        "rnc",
    ]

    # Menggunakan multiprocessing untuk menghitung nearest distance
    pool = Pool(processes=4)  # Jumlah proses sesuaikan dengan jumlah CPU yang tersedia
    results = []

    for row_umtscells in sheet_umtscells.iter_rows(min_row=2, values_only=True):
        result = pool.apply_async(
            calculate_nearest_distance, args=(row_umtscells, sheet_sites)
        )
        results.append(result)

    # Menunggu semua proses selesai
    pool.close()
    pool.join()

    # Mengambil hasil perhitungan nearest distance
    data_rows = []
    for result in results:
        data = result.get()
        data_rows.append(data)

    # Menyimpan hasil ke file CSV
    result_file_name = f"result_{datetime.now().strftime('%Y%m%d%H%M%S')}.csv"
    with open(result_file_name, "w", newline="") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(header)
        writer.writerows(data_rows)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python script.py <file_name>")
    else:
        file_name = sys.argv[1]
        main(file_name)
