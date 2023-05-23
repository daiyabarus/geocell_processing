import sys
from datetime import datetime
from multiprocessing import Manager, Pool

import openpyxl
from geopy.distance import geodesic


def calculate_distance(coords1, coords2):
    return geodesic(coords1, coords2).km


def find_nearest_distance(row, coords, manager):
    min_distance = float("inf")
    nearest_lc = None

    for lc, lat, lon in coords:
        if lc != row[0]:
            distance = calculate_distance((row[1], row[2]), (lat, lon))
            if distance > 0 and distance < min_distance:
                min_distance = distance
                nearest_lc = lc

    # If all distances are 0, find the nearest non-zero distance
    if min_distance == 0:
        for lc, lat, lon in coords:
            if lc != row[0]:
                distance = calculate_distance((row[1], row[2]), (lat, lon))
                if distance > 0:
                    min_distance = distance
                    nearest_lc = lc
                    break

    row[3] = min_distance
    row.append(nearest_lc)
    manager.append(row)


def process_rows(data, coords):
    with Manager() as manager:
        result = manager.list()

        pool = Pool()
        for row in data:
            pool.apply_async(find_nearest_distance, args=(row, coords, result))
        pool.close()
        pool.join()

        return list(result)


def main():
    if len(sys.argv) != 2:
        print("Usage: python main_pool.py input_file.xlsx")
        return

    filename = sys.argv[1]
    current_datetime = datetime.now().strftime("%Y%m%d%H%M")
    output_filename = f"Result_{current_datetime}.xlsx"

    try:
        wb = openpyxl.load_workbook(filename)
        sheet = wb.active
        data = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(list(row))

        coords = [(row[0], row[1], row[2]) for row in data]

        results = process_rows(data, coords)

        # Create a new workbook for the results
        result_wb = openpyxl.Workbook()
        result_sheet = result_wb.active

        # Write the results to the sheet
        result_sheet.append(["LC", "Latitude", "Longitude", "ISD", "1st Tier"])
        for row in results:
            result_sheet.append(row)

        # Save the result workbook
        result_wb.save(output_filename)
        print("Results saved to", output_filename)

    except FileNotFoundError:
        print("File not found!")


if __name__ == "__main__":
    main()
