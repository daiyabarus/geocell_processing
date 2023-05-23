import os
from typing import List

from openpyxl import load_workbook

from enumlist.enumlist import (
    referenceSheetNaming,
    referencesitescol,
    referenceumtscellscol,
)
from sheet.sheet_sites import sheet_Sites
from sheet.sheet_umtscells import sheet_UmtsCells


def getval(rawdata):
    return str(rawdata).strip().replace("None", "")


class collectreferencedata:
    def __init__(self, reference_filepath: str) -> None:
        self.reference_filepath = reference_filepath
        self.reference_sheetnaming = referenceSheetNaming()

    def collect_sheet_sites(self) -> list:
        list_of_data: List[sheet_Sites] = list()
        col = referencesitescol()

        if os.path.exists(self.reference_filepath):
            wb = load_workbook(self.reference_filepath, read_only=True, data_only=True)
            ws_name = self.reference_sheetnaming

            if ws_name in wb.sheetnames:
                ws = wb[ws_name]

                for row in ws.iter_rows(
                    min_row=col.ROW_START,
                    min_col=col.COL_START,
                    max_col=col.COL_END,
                    values_only=True,
                ):
                    lc = getval(row[col.sites_locCode])

                    if lc != "":
                        instance = referencesitescol(
                            sites_loccode=lc,
                            sitename=getval(row[col.siteName]),
                            siteisd=getval(row[col.siteIsd]),
                            parent=getval(row[col.parent]),
                            csd=getval(row[col.csd]),
                            cma=getval(row[col.cma]),
                            er=getval(row[col.er]),
                            prv=getval(row[col.prv]),
                            latitude=getval(row[col.latitude]),
                            longitude=getval(row[col.longitude]),
                            towertype=getval(row[col.towerType]),
                            rnc=getval(row[col.rnc]),
                            bsc=getval(row[col.rnc]),
                            bsc=getval(row[col.bsc]),
                            gsm=getval(row[col.GSM]),
                            umts=getval(row[col.UMTS]),
                            lte=getval(row[col.LTE]),
                            nr=getval(row[col.nr]),
                            nbiot=getval(row[col.NBIOT]),
                            csduid=getval(row[col.csdUid]),
                            mocnnodes=getval(row[col.mocnNodes]),
                            azimuthcount=getval(row[col.azimuthCount]),
                            antennaheight=getval(row[col.avgAntennaHeight]),
                            lte_shifted=getval(row[col.lte_shifted]),
                        )
                        list_of_data.append(instance)
        return list_of_data

    def collect_sheet_umtscells(self) -> list:
        list_of_data: list[sheet_UmtsCells] = list()
        col = referenceumtscellscol()

        if os.path.exists(self.reference_filepath):
            wb = load_workbook(self.reference_filepath, read_only=True, data_only=True)
            ws_name = self.reference_sheetnaming

            if ws_name in wb.sheetnames:
                ws = wb[ws_name]

                for row in ws.iter_rows(
                    min_row=col.ROW_START,
                    min_col=col.COL_START,
                    max_col=col.COL_END,
                    values_only=True,
                ):
                    lc = getval(row[col.locCode])

                    if lc != "":
                        instance = referenceumtscellscol(
                            nodeb=getval(row[col.nodeB]),
                            loccode=lc,
                            cellid=getval(row[col.cellId]),
                            sc=getval(row[col.sc]),
                            lac=getval(row[col.lac]),
                            rac=getval(row[col.rac]),
                            uarfcnul=getval(row[col.uarfcnUl]),
                            uarfcndl=getval(row[col.uarfcnDl]),
                            azimuth=getval(row[col.azimuth]),
                            hbw=getval(row[col.HBW]),
                            erpsectorstatusdesc=getval(row[col.erpSectorStatusDesc]),
                            frequency=getval(row[col.frequency]),
                        )
                        list_of_data.append(instance)
        return list_of_data
