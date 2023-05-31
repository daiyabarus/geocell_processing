import os

from openpyxl import load_workbook

from _enums_ import (  # UMTSCellsHeader,; UMTSNodesHeader, GSMCellsHeader,; GSMNodesHeader,; LTECellsHeader,;
    NrCellsHeader,
    RadioNodesHeader,
    SitesHeader,
)


class collectDataAll:
    def __init__(self, newReference_filepath) -> None:
        self.newReference_filepath = newReference_filepath
        self.dtref_nrcells = dict()
        # self.dtref_ltecells = dict()
        # self.dtref_umtscells = dict()
        # self.dtref_gsmcells = dict()
        self.dtref_radionodes = dict()
        # self.dtref_umtsnodes = dict()
        # self.dtref_gsmnodes = dict()
        self.dtref_sites = dict()
        self.collect_data()

    def collect_data(self):
        self.collect_ref_nrcells()
        # self.__coleect_ref_ltecells()
        # self.__collect_ref_umtscells()
        # self.__collect_ref_gsmcells()
        self.collect_ref_sites()
        self.collect_ref_radionodes()
        # self.__collect_ref_umtsnodes()
        # self.__collect_ref_gsmnodes()

    def collect_ref_nrcells(self):
        if os.path.exists(self.newReference_filepath):
            wb = load_workbook(
                self.newReference_filepath, read_only=True, data_only=True
            )
            ws = wb["nrCells"]

            for row in ws.iter_rows(
                min_row=NrCellsHeader.ROWSTART.value,
                min_col=NrCellsHeader.COLUMNSTART.value,
                max_col=NrCellsHeader.COLUMNEND.value,
                values_only=True,
            ):
                nRCellDUId = str(row[NrCellsHeader.NRCELLDUID.value]).strip()
                nodeId = str(row[NrCellsHeader.NODEID.value]).strip()
                locCode = str(row[NrCellsHeader.LOCCODE.value]).strip()
                gNbDuFunctionId = str(row[NrCellsHeader.GNBDUFUNCTIONID.value]).strip()
                administrativeState = str(
                    row[NrCellsHeader.ADMINISTRATIVESTATE.value]
                ).strip()
                availabilityStatus = str(
                    row[NrCellsHeader.AVAILABILITYSTATUS.value]
                ).strip()
                cellBarred = str(row[NrCellsHeader.CELLBARRED.value]).strip()
                cellLocalId = str(row[NrCellsHeader.CELLLOCALID.value]).strip()
                cellReservedForOperator = str(
                    row[NrCellsHeader.CELLRESERVEDFOROPERATOR.value]
                ).strip()
                cellState = str(row[NrCellsHeader.CELLSTATE.value]).strip()
                nCGI = str(row[NrCellsHeader.NCGI.value]).strip()
                nCI = str(row[NrCellsHeader.NCI.value]).strip()
                nRPCI = str(row[NrCellsHeader.NRPCI.value]).strip()
                nRSectorCarrier = str(row[NrCellsHeader.NRSECTORCARRIER.value]).strip()
                nRTAC = str(row[NrCellsHeader.NRTAC.value]).strip()
                operationalState = str(
                    row[NrCellsHeader.OPERATIONALSTATE.value]
                ).strip()
                qQualMin = str(row[NrCellsHeader.QQUALMIN.value]).strip()
                qRxLevMin = str(row[NrCellsHeader.QRXLEVMIN.value]).strip()
                rachRootSequence = str(
                    row[NrCellsHeader.RACHROOTSEQUENCE.value]
                ).strip()
                ssbDuration = str(row[NrCellsHeader.SSBDURATION.value]).strip()
                ssbFrequency = str(row[NrCellsHeader.SSBFREQUENCY.value]).strip()
                ssbFrequencyAutoSelected = str(
                    row[NrCellsHeader.SSBFREQUENCYAUTOSELECTED.value]
                ).strip()
                ssbOffset = str(row[NrCellsHeader.SSBOFFSET.value]).strip()
                ssbPeriodicity = str(row[NrCellsHeader.SSBPERIODICITY.value]).strip()
                ssbSubCarrierSpacing = str(
                    row[NrCellsHeader.SSBSUBCARRIERSPACING.value]
                ).strip()
                subCarrierSpacing = str(
                    row[NrCellsHeader.SUBCARRIERSPACING.value]
                ).strip()
                tddLteCoexistence = str(
                    row[NrCellsHeader.TDDLTECOEXISTENCE.value]
                ).strip()
                tddSpecialSlotPattern = str(
                    row[NrCellsHeader.TDDSPECIALSLOTPATTERN.value]
                ).strip()
                tddUlDlPattern = str(row[NrCellsHeader.TDDULDLPATTERN.value]).strip()
                arfcnDL = str(row[NrCellsHeader.ARFCNDL.value]).strip()
                arfcnUL = str(row[NrCellsHeader.ARFCNUL.value]).strip()
                bSChannelBwDL = str(row[NrCellsHeader.BSCHANNELBWDL.value]).strip()
                bSChannelBwUL = str(row[NrCellsHeader.BSCHANNELBWUL.value]).strip()
                azimuth = str(row[NrCellsHeader.AZIMUTH.value]).strip()
                erpSectorStatusDesc = str(
                    row[NrCellsHeader.ERPSECTORSTATUSDESC.value]
                ).strip()
                freq = str(row[NrCellsHeader.FREQ.value]).strip()
                bandList = str(row[NrCellsHeader.BANDLIST.value]).strip()
                pLMNIdList = str(row[NrCellsHeader.PLMNIDLIST.value]).strip()
                essScPairId = str(row[NrCellsHeader.ESSSCPAIRID.value]).strip()
                NRCellDU_userLabel = str(
                    row[NrCellsHeader.NRCELLDUUSERLABEL.value]
                ).strip()
                erpSectorTech = str(row[NrCellsHeader.ERPSECTORTECH.value]).strip()

                if locCode:
                    if not locCode in self.dtref_nrcells:
                        self.dtref_nrcells[locCode] = [
                            nRCellDUId,
                            nodeId,
                            locCode,
                            gNbDuFunctionId,
                            administrativeState,
                            availabilityStatus,
                            cellBarred,
                            cellLocalId,
                            cellReservedForOperator,
                            cellState,
                            nCGI,
                            nCGI,
                            nCI,
                            nRPCI,
                            nRSectorCarrier,
                            nRTAC,
                            operationalState,
                            qQualMin,
                            qRxLevMin,
                            rachRootSequence,
                            ssbDuration,
                            ssbFrequency,
                            ssbFrequencyAutoSelected,
                            ssbOffset,
                            ssbPeriodicity,
                            ssbSubCarrierSpacing,
                            subCarrierSpacing,
                            tddLteCoexistence,
                            tddSpecialSlotPattern,
                            tddUlDlPattern,
                            arfcnDL,
                            arfcnUL,
                            bSChannelBwDL,
                            bSChannelBwUL,
                            azimuth,
                            erpSectorStatusDesc,
                            freq,
                            bandList,
                            pLMNIdList,
                            essScPairId,
                            NRCellDU_userLabel,
                            erpSectorTech,
                        ]

    def collect_ref_sites(self):
        if os.path.exists(self.newReference_filepath):
            wb = load_workbook(
                self.newReference_filepath, read_only=True, data_only=True
            )
            ws = wb["sites"]

            for row in ws.iter_rows(
                min_row=SitesHeader.ROW_START.value,
                min_col=SitesHeader.COLUMN_START.value,
                max_col=SitesHeader.COLUMN_END.value,
                values_only=True,
            ):
                locCode_1 = str(row[SitesHeader.LOCCODE.value]).strip()
                sitename = str(row[SitesHeader.SITENAME.value]).strip()
                siteisd = str(row[SitesHeader.SITEISD.value]).strip()
                parent = str(row[SitesHeader.PARENT.value]).strip()
                csd = str(row[SitesHeader.CSD.value]).strip()
                cma = str(row[SitesHeader.CMA.value]).strip()
                er = str(row[SitesHeader.ER.value]).strip()
                prv = str(row[SitesHeader.PRV.value]).strip()
                latitude = str(row[SitesHeader.LATITUDE.value]).strip()
                longitude = str(row[SitesHeader.LONGITUDE.value]).strip()
                towertype = str(row[SitesHeader.TOWER_TYPE.value]).strip()
                rnc = str(row[SitesHeader.RNC.value]).strip()
                bsc = str(row[SitesHeader.BSC.value]).strip()
                gsm = str(row[SitesHeader.GSM.value]).strip()
                umts = str(row[SitesHeader.UMTS.value]).strip()
                lte = str(row[SitesHeader.LTE.value]).strip()
                nr = str(row[SitesHeader.NR.value]).strip()
                nbiot = str(row[SitesHeader.NBIOT.value]).strip()
                csduid = str(row[SitesHeader.CSDUID.value]).strip()
                mocnnodes = str(row[SitesHeader.MOCNNODES.value]).strip()
                azimuthcount = str(row[SitesHeader.AZIMUTHCOUNT.value]).strip()
                avgantennaheight = str(row[SitesHeader.AVGANTENNAHEIGHT.value]).strip()
                lte_shifted = str(row[SitesHeader.LTESHIFTED.value]).strip()

                if locCode_1:
                    if not locCode_1 in self.dtref_sites:
                        self.dtref_sites[locCode_1] = [
                            locCode_1,
                            sitename,
                            siteisd,
                            parent,
                            csd,
                            cma,
                            er,
                            prv,
                            latitude,
                            longitude,
                            towertype,
                            rnc,
                            bsc,
                            gsm,
                            umts,
                            lte,
                            nr,
                            nbiot,
                            csduid,
                            mocnnodes,
                            azimuthcount,
                            avgantennaheight,
                            lte_shifted,
                        ]

    def collect_ref_radionodes(self):
        if os.path.exists(self.newReference_filepath):
            wb = load_workbook(
                self.newReference_filepath, read_only=True, data_only=True
            )
            ws = wb["radioNodes"]

            for row in ws.iter_rows(
                min_row=RadioNodesHeader.ROW_START.value,
                min_col=RadioNodesHeader.COLUMN_START.value,
                max_col=RadioNodesHeader.COLUMN_END.value,
                values_only=True,
            ):
                nodeid = str(row[RadioNodesHeader.NODEID.value]).strip()
                enbid = str(row[RadioNodesHeader.ENBID.value]).strip()
                loccode_2 = str(row[RadioNodesHeader.LOCCODE.value]).strip()
                subnetwork = str(row[RadioNodesHeader.SUBNETWORK.value]).strip()
                system = str(row[RadioNodesHeader.SYSTEM.value]).strip()
                swversionid = str(row[RadioNodesHeader.SWVERSIONID.value]).strip()
                swrelease = str(row[RadioNodesHeader.SWRELEASE.value]).strip()
                vendor = str(row[RadioNodesHeader.VENDOR.value]).strip()
                mocn = str(row[RadioNodesHeader.MOCN.value]).strip()
                gnbid = str(row[RadioNodesHeader.GNBID.value]).strip()
                hwtype = str(row[RadioNodesHeader.HWTYPE.value]).strip()

                if loccode_2:
                    if not loccode_2 in self.dtref_radionodes:
                        self.dtref_radionodes[loccode_2] = [
                            nodeid,
                            enbid,
                            loccode_2,
                            subnetwork,
                            system,
                            swversionid,
                            swrelease,
                            vendor,
                            mocn,
                            gnbid,
                            hwtype,
                        ]
