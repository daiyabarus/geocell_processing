import csv

from openpyxl import load_workbook

from enums import NrCellsHeader, RadioNodesHeader, SitesHeader

# Load workbook in read-only mode
wb = load_workbook("newReference.xlsx", read_only=True)

# Load nrCells sheet
nr_cells_sheet = wb["nrCells"]
nr_cells_data = [
    tuple(row) for row in nr_cells_sheet.iter_rows(min_row=2, values_only=True)
]

# Load radioNodes sheet
radio_nodes_sheet = wb["radioNodes"]
radio_nodes_data = [
    tuple(row) for row in radio_nodes_sheet.iter_rows(min_row=2, values_only=True)
]

# Load sites sheet
sites_sheet = wb["sites"]
sites_data = [tuple(row) for row in sites_sheet.iter_rows(min_row=2, values_only=True)]

# Close workbook
wb.close()

# Load netstatRemedy data into a dictionary
netstat_remedy_data = {}
with open("netstatRemedy.csv", "r") as csvfile:
    reader = csv.DictReader(csvfile, delimiter=";")
    for row in reader:
        erp_sector_name = row["ERP Sector Name"]
        netstat_remedy_data[erp_sector_name] = (
            row["ERP Site Name"],
            row["ERP Sector Elect Downtilt"],
            row["ERP Sector Mechanical Downtilt"],
            row["ERP Sector Antenna Omnidirectional"],
        )

radio_nodes_loc_codes = set(
    radio_nodes_row[RadioNodesHeader.locCode.value - 1]
    for radio_nodes_row in radio_nodes_data
)
sites_loc_codes = set(
    site_row[SitesHeader.locCode.value - 1] for site_row in sites_data
)

# Perform the join operation
joined_data = []
for nr_cells_row in nr_cells_data:
    loc_code = nr_cells_row[NrCellsHeader.locCode.value - 1]
    if loc_code in radio_nodes_loc_codes and loc_code in sites_loc_codes:
        matching_radio_nodes = [
            rn_row
            for rn_row in radio_nodes_data
            if rn_row[RadioNodesHeader.locCode.value - 1] == loc_code
        ]
        matching_sites = [
            site_row
            for site_row in sites_data
            if site_row[SitesHeader.locCode.value - 1] == loc_code
        ]
        erp_sector_name = nr_cells_row[NrCellsHeader.nRCellDUId.value - 1]
        if erp_sector_name in netstat_remedy_data:
            nr_row = netstat_remedy_data[erp_sector_name]
            for rn_row in matching_radio_nodes:
                for site_row in matching_sites:
                    joined_row = nr_cells_row + rn_row + site_row + nr_row
                    joined_data.append(joined_row)

# Write joined data to output.csv
with open("output.csv", "w", newline="") as csvfile:
    writer = csv.writer(csvfile)
    writer.writerow(
        [header.name for header in NrCellsHeader]
        + [header.name for header in RadioNodesHeader]
        + [header.name for header in SitesHeader]
        + [
            "ERP Site Name",
            "ERP Sector Elect Downtilt",
            "ERP Sector Mechanical Downtilt",
            "ERP Sector Antenna Omnidirectional",
        ]
    )
    writer.writerows(joined_data)
