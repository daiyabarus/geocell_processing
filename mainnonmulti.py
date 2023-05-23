# import sys

# from geopy.distance import geodesic
# from openpyxl import load_workbook


# def find_nearest_distance(filename):
#     # Membaca file Excel
#     workbook = load_workbook(filename)
#     sheet = workbook.active

#     # Mencari indeks kolom Latitude, Longitude, dan Nearest_Distance
#     headers = [cell.value for cell in sheet[1]]
#     lat_index = headers.index("Latitude") + 1
#     long_index = headers.index("Longitude") + 1
#     nearest_dist_index = headers.index("Nearest_Distance") + 1

#     # Mengiterasi setiap baris
#     for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
#         point1 = (row[lat_index - 1], row[long_index - 1])
#         nearest_distance = float(
#             "inf"
#         )  # Menginisialisasi jarak terdekat dengan nilai tak terhingga

#         # Membandingkan dengan setiap baris lainnya
#         for j, other_row in enumerate(
#             sheet.iter_rows(min_row=2, values_only=True), start=2
#         ):
#             if i != j:  # Tidak membandingkan dengan dirinya sendiri
#                 point2 = (other_row[lat_index - 1], other_row[long_index - 1])
#                 distance = geodesic(point1, point2).kilometers

#                 if distance < nearest_distance:
#                     nearest_distance = distance

#         # Memperbarui nilai jarak terdekat pada kolom Nearest_Distance
#         sheet.cell(row=i, column=nearest_dist_index, value=nearest_distance)

#     # Menyimpan perubahan pada file Excel
#     workbook.save(filename)


# if __name__ == "__main__":
#     if len(sys.argv) < 2:
#         print("Usage: python program.py <filename.xlsx>")
#         sys.exit(1)

#     filename = sys.argv[1]
#     find_nearest_distance(filename)
#     print("Nearest distances have been calculated and updated in the Excel file.")

# import sys
# from multiprocessing import Pool

# from geopy.distance import geodesic
# from openpyxl import load_workbook


# def calculate_nearest_distance(row, data, lat_index, long_index):
#     point1 = (row[lat_index - 1], row[long_index - 1])
#     nearest_distance = float("inf")

#     for other_row in data:
#         point2 = (other_row[lat_index - 1], other_row[long_index - 1])
#         distance = geodesic(point1, point2).kilometers

#         if distance < nearest_distance:
#             nearest_distance = distance

#     return nearest_distance


# def find_nearest_distance(filename):
#     workbook = load_workbook(filename)
#     sheet = workbook.active
#     data = list(sheet.iter_rows(min_row=2, values_only=True))
#     headers = [cell.value for cell in sheet[1]]
#     lat_index = headers.index("Latitude") + 1
#     long_index = headers.index("Longitude") + 1
#     nearest_dist_index = headers.index("Nearest_Distance") + 1

#     pool = Pool()

#     results = []
#     for row in data:
#         results.append(
#             pool.apply_async(
#                 calculate_nearest_distance, args=(row, data, lat_index, long_index)
#             )
#         )

#     pool.close()
#     pool.join()

#     for i, result in enumerate(results, start=2):
#         nearest_distance = result.get()
#         sheet.cell(row=i, column=nearest_dist_index, value=nearest_distance)

#     workbook.save(filename)


# if __name__ == "__main__":
#     if len(sys.argv) < 2:
#         print("Usage: python program.py <filename.xlsx>")
#         sys.exit(1)

#     filename = sys.argv[1]
#     find_nearest_distance(filename)
#     print("Nearest distances have been calculated and updated in the Excel file.")

# import sys
# from multiprocessing import Pool

# from geopy.distance import geodesic
# from openpyxl import load_workbook


# def calculate_nearest_distance(row, data, lat_index, long_index):
#     point1 = (row[lat_index - 1], row[long_index - 1])
#     nearest_distance = float("inf")

#     for other_row in data:
#         point2 = (other_row[lat_index - 1], other_row[long_index - 1])
#         distance = geodesic(point1, point2).kilometers

#         if distance < nearest_distance:
#             nearest_distance = distance

#     return nearest_distance


# def find_nearest_distance(filename):
#     workbook = load_workbook(filename)
#     sheet = workbook.active
#     data = list(sheet.iter_rows(min_row=2, values_only=True))
#     headers = [cell.value for cell in sheet[1]]
#     lat_index = headers.index("Latitude") + 1
#     long_index = headers.index("Longitude") + 1
#     nearest_dist_index = headers.index("Nearest_Distance") + 1

#     pool = Pool()  # Menggunakan multiprocessing pool

#     results = []
#     for row in data:
#         results.append(
#             pool.apply_async(
#                 calculate_nearest_distance, args=(row, data, lat_index, long_index)
#             )
#         )

#     pool.close()
#     pool.join()

#     for i, result in enumerate(results, start=2):
#         nearest_distance = result.get()
#         sheet.cell(row=i, column=nearest_dist_index, value=nearest_distance)

#     workbook.save(filename)


# if __name__ == "__main__":
#     if len(sys.argv) < 2:
#         print("Usage: python program.py <filename.xlsx>")
#         sys.exit(1)

#     filename = sys.argv[1]
#     find_nearest_distance(filename)
#     print("Nearest distances have been calculated and updated in the Excel file.")

# # # # # # OK
# import sys
# from multiprocessing import Manager, Pool

# import openpyxl
# from geopy.distance import geodesic


# def calculate_distance(coords1, coords2):
#     return geodesic(coords1, coords2).km


# def find_nearest_distance(row, coords, manager):
#     min_distance = float("inf")
#     nearest_lc = None

#     for lc, lat, lon in coords:
#         if lc != row[0]:
#             distance = calculate_distance((row[1], row[2]), (lat, lon))
#             if distance < min_distance:
#                 min_distance = distance
#                 nearest_lc = lc

#     row[3] = min_distance
#     row.append(nearest_lc)
#     manager.append(row)


# def process_rows(data, coords):
#     with Manager() as manager:
#         result = manager.list()

#         pool = Pool()
#         for row in data:
#             pool.apply_async(find_nearest_distance, args=(row, coords, result))
#         pool.close()
#         pool.join()

#         return list(result)


# def main():
#     if len(sys.argv) != 2:
#         print("Usage: python program.py input_file.xlsx")
#         return

#     filename = sys.argv[1]

#     try:
#         wb = openpyxl.load_workbook(filename)
#         sheet = wb.active
#         data = []

#         for row in sheet.iter_rows(min_row=2, values_only=True):
#             data.append(list(row))

#         coords = [(row[0], row[1], row[2]) for row in data]

#         results = process_rows(data, coords)

#         for row in results:
#             print(row)

#     except FileNotFoundError:
#         print("File not found!")


# if __name__ == "__main__":
#     main()

# import sys
# from multiprocessing import Manager, Pool

# import openpyxl
# from geopy.distance import geodesic


# def calculate_distance(coords1, coords2):
#     return geodesic(coords1, coords2).km


# def find_nearest_distance(row, coords, manager):
#     min_distance = float("inf")
#     nearest_lc = None

#     for lc, lat, lon in coords:
#         if lc != row[0]:
#             distance = calculate_distance((row[1], row[2]), (lat, lon))
#             if distance < min_distance:
#                 min_distance = distance
#                 nearest_lc = lc

#     row[3] = min_distance
#     row.append(nearest_lc)
#     manager.append(row)


# def process_rows(data, coords):
#     with Manager() as manager:
#         result = manager.list()

#         pool = Pool()
#         for row in data:
#             pool.apply_async(find_nearest_distance, args=(row, coords, result))
#         pool.close()
#         pool.join()

#         return list(result)


# def main():
#     if len(sys.argv) != 2:
#         print("Usage: python program.py input_file.xlsx")
#         return

#     filename = sys.argv[1]
#     output_filename = "Result_tanggal.xlsx"

#     try:
#         wb = openpyxl.load_workbook(filename)
#         sheet = wb.active
#         data = []

#         for row in sheet.iter_rows(min_row=2, values_only=True):
#             data.append(list(row))

#         coords = [(row[0], row[1], row[2]) for row in data]

#         results = process_rows(data, coords)

#         # Create a new workbook for the results
#         result_wb = openpyxl.Workbook()
#         result_sheet = result_wb.active

#         # Write the results to the sheet
#         result_sheet.append(
#             ["LC", "Latitude", "Longitude", "Nearest_Distance", "Nearest_LC"]
#         )
#         for row in results:
#             result_sheet.append(row)

#         # Save the result workbook
#         result_wb.save(output_filename)
#         print("Results saved to", output_filename)

#     except FileNotFoundError:
#         print("File not found!")


# if __name__ == "__main__":
#     main()


import sys
from datetime import datetime
from multiprocessing import Manager, Pool

import openpyxl
from geopy.distance import geodesic


def calculate_distance(coords1, coords2):
    return geodesic(coords1, coords2).km


def find_nearest_distance(row, coords, manager):
    min_distance = float("inf")
    nearest_lc = None

    for lc, lat, lon in coords:
        if lc != row[0]:
            distance = calculate_distance((row[1], row[2]), (lat, lon))
            if distance < min_distance:
                min_distance = distance
                nearest_lc = lc

    row[3] = min_distance
    row.append(nearest_lc)
    manager.append(row)


def process_rows(data, coords):
    with Manager() as manager:
        result = manager.list()

        pool = Pool()
        for row in data:
            pool.apply_async(find_nearest_distance, args=(row, coords, result))
        pool.close()
        pool.join()

        return list(result)


def main():
    if len(sys.argv) != 2:
        print("Usage: python program.py input_file.xlsx")
        return

    filename = sys.argv[1]
    current_datetime = datetime.now().strftime("%Y%m%d_%H%M")
    output_filename = f"ISD_{current_datetime}.xlsx"

    try:
        wb = openpyxl.load_workbook(filename)
        sheet = wb.active
        data = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(list(row))

        coords = [(row[0], row[1], row[2]) for row in data]

        results = process_rows(data, coords)

        # Create a new workbook for the results
        result_wb = openpyxl.Workbook()
        result_sheet = result_wb.active

        # Write the results to the sheet
        result_sheet.append(["LC", "Latitude", "Longitude", "ISD", "1st Tier"])
        for row in results:
            result_sheet.append(row)

        # Save the result workbook
        result_wb.save(output_filename)
        print("Results saved to", output_filename)

    except FileNotFoundError:
        print("File not found!")


if __name__ == "__main__":
    main()
